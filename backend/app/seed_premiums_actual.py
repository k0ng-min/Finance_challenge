"""보험사 다이렉트 보험료 자료와 값별 provenance를 적재한다.

2026-08-19 이전에는 보험다모아 비교공시(crawl_premiums.py → seed_premiums.py)를 썼다.
그건 "표준조건" 한 값이라 보험사가 실제로 파는 여러 등급(플랜)의 가격 차이를 보여주지
못했다. 이 스크립트는 그걸 완전히 대체한다 — 사용자가 각 사 다이렉트 계산기에서
나이·성별·등급별 자료(backend/data/source_files/insurer_premiums_*.xlsx)를 적재한다.
대부분은 직접조회값이지만 카카오는 3일 값을 1일로 환산했고, 메리츠 추천플랜은 양쪽
직접조회 플랜의 산술평균이다. 이 차이를 value_origin과 변환 메타데이터로 보존한다.

시트 구조가 보험사마다 다르다(카카오는 나이·성별·등급이 한 행씩인 세로형, 나머지는
등급 3개가 열로 나란한 가로형) — _rows_from_sheet()가 그 차이를 흡수해서 전부
ParsedPremium으로 통일한다.

아직 실제 가격을 조회하지 못한 보험사는 시트 자체가 없다. 그 보험사는 이 스크립트를
몇 번을 돌려도 그냥 조용히 건너뛴다 — 나중에 같은 형식의 시트를 엑셀에 추가하고
_SHEET_CONFIG·_BASIS에 한 줄씩 더하면 채워진다(DB손해보험이 2026-08-25에 그렇게 들어왔다).

Run from ``backend``::

    python -m app.seed_premiums_actual
"""
from __future__ import annotations

import sys
from dataclasses import dataclass, replace
from datetime import date
from pathlib import Path

import openpyxl

from app.database import SessionLocal, engine
from app import models  # noqa: F401  (모델 등록)
from app import schema_migrations
from app.models.kb import (
    PREMIUM_ORIGIN_DERIVED,
    PREMIUM_ORIGIN_DIRECT_QUOTE,
    PREMIUM_ORIGIN_IMPUTED,
    Insurer,
    InsurerPremium,
)


def _ensure_new_schema() -> None:
    """앱 기동과 같은 멱등 migration을 적용한다."""
    schema_migrations.apply(engine)

_SOURCE_DIR = Path(__file__).resolve().parents[1] / "data" / "source_files"
DEFAULT_PATH = _SOURCE_DIR / "insurer_premiums_2026-08.xlsx"

#: 나중에 따로 받은 보험사는 그 파일을 그대로 둔다. 통합 엑셀에 시트를 옮겨 붙이려면
#: openpyxl로 다시 저장해야 하는데, 그 파일의 카카오 시트에는 1일 환산 보험료가
#: `=ROUND(D3/3,0)` 수식으로 들어 있다(486칸). openpyxl은 저장할 때 수식의 **계산된 값**을
#: 버리므로, 다시 저장하는 순간 카카오 가격이 통째로 빈 값이 된다 — 시트를 합치는 대신
#: 파일을 나란히 읽는다.
EXTRA_PATHS: tuple[Path, ...] = (
    _SOURCE_DIR / "insurer_premiums_shinhan_2026-08.xlsx",
)

_SEX_MAP = {"남": "M", "여": "F"}

#: 시트 이름 -> (보험사 코드, 세로형 여부, 화면에 대표로 보여줄 표준 등급명)
_SHEET_CONFIG: dict[str, tuple[str, bool, str]] = {
    "카카오": ("KAKAOPAY", True, "베이직"),
    "현대해상": ("HYUNDAI", False, "표준형"),
    "kb": ("KB", False, "표준형"),
    "삼성": ("SAMSUNG", False, "표준플랜"),
    # 메리츠는 실속플랜·추천플랜·보장이큰플랜 세 등급이고 추천플랜이 표준 자리다.
    "메리츠": ("MERITZ", False, "추천플랜"),
    "db": ("DB", False, "표준형"),
    # 신한EZ손보는 실속케어·안심케어 두 등급만 판다(고급 자리가 없다 — insurer_tiers.py
    # 참고). 그래서 표준 자리는 두 번째인 안심케어다.
    "신한": ("SHINHAN", False, "안심케어"),
}

#: 가격표 시트의 등급명 열 헤더가 담보 가입금액표(InsurerPlanCoverage)와 다르게 적힌
#: 경우를 여기서 맞춘다 — 현대해상은 가격표엔 "실속(원)"처럼 "형"이 빠져 있는데
#: 담보한도표(seed_plan_coverage.py)는 "실속형"으로 적혀 있다(둘 다 사용자가 각 사
#: 사이트에서 그대로 옮긴 값이라 시트마다 표기가 달랐다). 두 테이블을 등급명으로
#: 이어 붙여 써야 하므로(PlanCoverageBoard) 여기서 하나로 통일한다.
_PLAN_NAME_ALIASES: dict[str, dict[str, str]] = {
    "HYUNDAI": {"실속": "실속형", "표준": "표준형", "고급": "고급형"},
    # 가격 시트는 "보장이 큰 플랜"(띄어쓰기), 보장금액표·등급 매핑은 "보장이큰플랜"이다.
    # 여기서 맞추지 않으면 등급으로 가격을 못 찾아 메리츠만 조용히 값이 안 뜬다.
    # 가운데 등급도 마찬가지다 — 가격 시트 헤더는 "표준형", 보장금액표는 "추천플랜"이다.
    # (참고: 이 표준형 열 122행은 전부 실속플랜과 보장이큰플랜의 산술평균이다. 다이렉트에서
    #  직접 조회한 값이 아니라 두 등급에서 뽑아낸 값이라는 걸 알고 쓰는 것으로 정리됐다.)
    "MERITZ": {"보장이 큰 플랜": "보장이큰플랜", "표준형": "추천플랜"},
}

#: 조회 조건 — 보험사마다 산출 전제가 달라 각각 남긴다(근거 없이 숫자만 보여주지 않는다).
_BASIS: dict[str, str] = {
    "KAKAOPAY": "다이렉트 사이트 조회, 3일 보험료를 3으로 나눈 1일 환산가, 단독가입 기준",
    "HYUNDAI": "다이렉트 사이트 조회, 1일(24시간) 기준, 단독가입",
    "KB": "다이렉트 사이트 조회, 1일(24시간) 기준, 단독가입(만19세~만90세 가입 가능)",
    "SAMSUNG": "다이렉트 사이트 조회, 1일(24시간) 기준, 단독가입(항공지연 지수형 특약 2종 기본 포함)",
    "MERITZ": "다이렉트 사이트 조회, 1일(24시간) 기준, 단독가입(만19세~만79세 조회 가능)",
    "DB": "다이렉트 사이트 조회, 1일(24시간) 기준, 본인 단독가입(만19세~만79세 조회 가능)",
    "SHINHAN": "다이렉트 사이트 조회, 1일(24시간) 기준, 단독가입(만19세~만79세 가입 가능)",
}
_SOURCE = "보험사 다이렉트 홈페이지 보험료 계산기(직접 조회)"
_COLLECTED_AT = date(2026, 8, 17)
#: 나중에 따로 조회한 보험사는 조회일이 다르다 — 언제 본 값인지 화면에 그대로 밝히므로
#: 한 날짜로 뭉뚱그리지 않는다. DB손보는 시트 안내문에 "조회일: 2026-08-18 ~ 2026-08-23"으로
#: 적혀 있어, 그 구간의 마지막 날을 기준일로 둔다.
_COLLECTED_AT_BY_CODE: dict[str, date] = {
    "DB": date(2026, 8, 23),
    # 신한 시트에는 조회일 주석이 없다. 지어내지 않고 자료를 건네받은 날을 기준일로 둔다.
    "SHINHAN": date(2026, 8, 25),
}


@dataclass(frozen=True)
class ParsedPremium:
    age: int
    sex: str
    plan_name: str
    premium: int
    value_origin: str
    source_value: int | None
    source_period_days: int | None
    transformation: str | None
    transformation_reason: str | None
    source_reference: str


def _find_header_row(rows: list[tuple]) -> int:
    """'성별' 열이 있는 행을 헤더로 본다 — 시트마다 그 위에 제목/주의문구 행 수가 다르다."""
    for i, row in enumerate(rows):
        if len(row) > 1 and row[1] == "성별":
            return i
    raise ValueError("헤더 행('성별' 열)을 찾지 못했습니다.")


def _rows_from_sheet(
    sheet_name: str,
    rows: list[tuple],
    vertical: bool,
    source_filename: str,
) -> list[ParsedPremium]:
    """시트 값과 원본 셀을 provenance가 포함된 행으로 통일한다."""
    header_idx = _find_header_row(rows)
    header = rows[header_idx]
    data_rows = rows[header_idx + 1:]
    out: list[ParsedPremium] = []

    if vertical:
        # 카카오: 나이, 성별, 가입플랜, 3일보험료, 1일환산보험료 — 한 행에 등급 하나.
        for excel_row, row in enumerate(data_rows, start=header_idx + 2):
            if not row or not isinstance(row[0], (int, float)) or row[1] not in _SEX_MAP:
                continue
            age, sex_raw, plan_name, three_day, one_day = row[0], row[1], row[2], row[3], row[4]
            if three_day is None or one_day is None:
                continue
            out.append(ParsedPremium(
                age=int(age), sex=_SEX_MAP[sex_raw], plan_name=str(plan_name),
                premium=int(one_day), value_origin=PREMIUM_ORIGIN_DERIVED,
                source_value=int(three_day), source_period_days=3,
                transformation="ROUND(source_value / source_period_days, 0)",
                transformation_reason="3일 직접 조회값을 1일 비교용 값으로 환산",
                source_reference=f"{source_filename}#{sheet_name}!D{excel_row}:E{excel_row}",
            ))
        return out

    # 현대해상/kb/삼성: 나이, 성별, 등급1, 등급2, 등급3[, 비고] — 등급이 열로 나란함.
    plan_names = [str(h).split("(")[0] for h in header[2:] if h and str(h) != "비고"]
    for excel_row, row in enumerate(data_rows, start=header_idx + 2):
        if not row or not isinstance(row[0], (int, float)) or row[1] not in _SEX_MAP:
            continue
        age, sex_raw = row[0], row[1]
        for col_offset, plan_name in enumerate(plan_names):
            premium = row[2 + col_offset]
            if premium is None:
                continue  # 가입연령 범위 밖 등 — 이 나이·등급은 조회 자체가 안 됨
            column_number = 3 + col_offset
            column_letter = openpyxl.utils.get_column_letter(column_number)
            is_meritz_interpolation = sheet_name == "메리츠" and plan_name == "표준형"
            if is_meritz_interpolation:
                origin = PREMIUM_ORIGIN_IMPUTED
                source_value = None
                transformation = "ROUND((실속플랜 + 보장이큰플랜) / 2, 0)"
                reason = "두 직접 조회 플랜 사이의 산술평균으로 만든 중간값"
                reference = (
                    f"{source_filename}#{sheet_name}!C{excel_row},E{excel_row}->D{excel_row}"
                )
            else:
                origin = PREMIUM_ORIGIN_DIRECT_QUOTE
                source_value = int(premium)
                transformation = None
                reason = None
                reference = f"{source_filename}#{sheet_name}!{column_letter}{excel_row}"
            out.append(ParsedPremium(
                age=int(age), sex=_SEX_MAP[sex_raw], plan_name=plan_name,
                premium=int(premium), value_origin=origin,
                source_value=source_value, source_period_days=1,
                transformation=transformation, transformation_reason=reason,
                source_reference=reference,
            ))
    return out


def run(path: Path = DEFAULT_PATH) -> dict[str, int]:
    if not path.exists():
        raise SystemExit(f"{path} 가 없습니다. 실제 보험료 엑셀을 이 경로에 둔 뒤 다시 실행하세요.")

    _ensure_new_schema()

    # read_only로 연다 — 엑셀에서 다시 저장된 파일은 서식 정의가 어긋나 있을 때가
    # 있는데(실제로 메리츠 시트가 추가된 판본이 그랬다), 일반 모드는 그 서식을 읽다가
    # 통째로 실패한다. 값만 읽으면 되므로 서식을 건너뛴다.
    workbooks = [(path, openpyxl.load_workbook(path, data_only=True, read_only=True))]
    workbooks += [
        (extra, openpyxl.load_workbook(extra, data_only=True, read_only=True))
        for extra in EXTRA_PATHS if extra.exists()
    ]
    db = SessionLocal()
    try:
        code_to_id = {i.code: i.insurer_id for i in db.query(Insurer).all()}
        counts: dict[str, int] = {}

        for sheet_name, (insurer_code, vertical, standard_plan) in _SHEET_CONFIG.items():
            workbook = next(((p, w) for p, w in workbooks if sheet_name in w.sheetnames), None)
            if workbook is None:
                continue
            workbook_path, wb = workbook
            insurer_id = code_to_id.get(insurer_code)
            if insurer_id is None:
                continue

            # 이 보험사분은 통째로 새로 채운다(등급이 늘거나 줄 수 있어 부분 갱신보다
            # 안전하다) — 사용자가 "원래 있던 가격 싹 다 지우고 이걸로 바꿔달라"고
            # 명시적으로 요청한 부분이다.
            db.query(InsurerPremium).filter(InsurerPremium.insurer_id == insurer_id).delete()

            rows = list(wb[sheet_name].iter_rows(values_only=True))
            parsed = _rows_from_sheet(sheet_name, rows, vertical, workbook_path.name)
            aliases = _PLAN_NAME_ALIASES.get(insurer_code, {})
            parsed = [
                replace(row, plan_name=aliases.get(row.plan_name, row.plan_name))
                for row in parsed
            ]
            collected_at = _COLLECTED_AT_BY_CODE.get(insurer_code, _COLLECTED_AT)
            for row in parsed:
                db.add(InsurerPremium(
                    insurer_id=insurer_id, sex=row.sex, age=row.age, plan_name=row.plan_name,
                    is_standard_tier=(row.plan_name == standard_plan),
                    premium=row.premium, period_days=1,
                    value_origin=row.value_origin,
                    source_value=row.source_value,
                    source_period_days=row.source_period_days,
                    transformation=row.transformation,
                    transformation_reason=row.transformation_reason,
                    source_reference=row.source_reference,
                    product_name=row.plan_name,
                    age_range=None,
                    basis=_BASIS[insurer_code], source=_SOURCE, source_url=None,
                    collected_at=collected_at,
                ))
            counts[insurer_code] = len(parsed)

        db.commit()
        for code, n in counts.items():
            print(f"{code}: {n}건 적재")
        all_codes = {i.code for i in db.query(Insurer).all()}
        missing = sorted(all_codes - set(counts))
        if missing:
            print(f"이 엑셀에 시트가 없어 건너뜀(아직 가격 미확보): {', '.join(missing)}")
        return counts
    finally:
        db.close()
        for _path, workbook in workbooks:
            workbook.close()


if __name__ == "__main__":
    sys.exit(0 if run(Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_PATH) else 1)
