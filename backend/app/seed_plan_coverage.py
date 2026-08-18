"""보험사 다이렉트 사이트에서 사용자가 직접 조회한, 플랜(등급)별 담보 가입금액표를 적재한다.

InsurerPremium(가격)과 짝을 이루는 자료지만 성격이 다르다 — 이건 "얼마 내는지"가 아니라
"그 돈으로 뭘 얼마나 받는지"다. 6개사 전부(DB·메리츠 포함) 담보한도표를 갖고 있다 —
가격표(seed_premiums_actual.py)와 달리 DB·메리츠도 이번에 같이 채운다.

시트 구조(backend/data/source_files/insurer_plan_coverage_2026-08.xlsx의 "*_상세" 6개
시트)는 전부 동일하다:

    ('담보명', 등급1, 등급2, 등급3)
    (담보명, 금액1, 금액2, 금액3)
    ...
    (None, None, None, None)          <- 데이터 끝
    ('단위: 만원(...). 출처: ...')      <- 출처·주의사항 각주

금액 칸에 숫자가 아니라 "-"(미가입)·"미제공"(그 상품 자체에 없음)·"가입"(정액·한도
비공개) 같은 표시가 섞여 있어서 amount_text는 원문 표기를 그대로 문자열로 저장한다.

Run from ``backend``::

    python -m app.seed_plan_coverage
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import openpyxl

from app.database import Base, SessionLocal, engine
from app import models  # noqa: F401  (모델 등록)
from app.models.kb import Insurer, InsurerPlanCoverage

DEFAULT_PATH = Path(__file__).resolve().parents[1] / "data" / "source_files" / "insurer_plan_coverage_2026-08.xlsx"

#: 시트 이름 접두어 -> 보험사 코드
_SHEET_PREFIX_TO_CODE: dict[str, str] = {
    "카카오": "KAKAOPAY",
    "현대": "HYUNDAI",
    "KB": "KB",
    "삼성": "SAMSUNG",
    "DB": "DB",
    "메리츠": "MERITZ",
}
_SOURCE = "보험사 다이렉트 홈페이지 보험료 계산 화면(직접 조회)"
_COLLECTED_AT = date(2026, 8, 17)


def _parse_sheet(rows: list[tuple]) -> tuple[list[str], list[tuple[str, list]], str]:
    """(등급명 3개, [(담보명, [금액1,금액2,금액3]), ...], 출처 각주) 로 나눈다."""
    header = rows[0]
    plan_names = [str(h) for h in header[1:] if h]

    data: list[tuple[str, list]] = []
    footnote_lines: list[str] = []
    seen_blank = False
    for row in rows[1:]:
        if row[0] is None:
            seen_blank = True
            continue
        if seen_blank:
            footnote_lines.append(str(row[0]))
            continue
        label = str(row[0])
        amounts = list(row[1:1 + len(plan_names)])
        data.append((label, amounts))

    return plan_names, data, " ".join(footnote_lines)


def run(path: Path = DEFAULT_PATH) -> dict[str, int]:
    if not path.exists():
        raise SystemExit(f"{path} 가 없습니다. 담보 가입금액표 엑셀을 이 경로에 둔 뒤 다시 실행하세요.")

    Base.metadata.create_all(bind=engine)

    wb = openpyxl.load_workbook(path, data_only=True)
    db = SessionLocal()
    try:
        code_to_id = {i.code: i.insurer_id for i in db.query(Insurer).all()}
        counts: dict[str, int] = {}

        for sheet_name in wb.sheetnames:
            if not sheet_name.endswith("_상세"):
                continue
            prefix = sheet_name[: -len("_상세")]
            insurer_code = _SHEET_PREFIX_TO_CODE.get(prefix)
            if insurer_code is None:
                continue
            insurer_id = code_to_id.get(insurer_code)
            if insurer_id is None:
                continue

            rows = list(wb[sheet_name].iter_rows(values_only=True))
            plan_names, data, footnote = _parse_sheet(rows)

            # 통째로 새로 채운다 — 담보 항목 자체가 늘거나 줄 수 있어 부분 갱신보다 안전하다.
            db.query(InsurerPlanCoverage).filter(InsurerPlanCoverage.insurer_id == insurer_id).delete()

            sort_order = 0
            for label, amounts in data:
                for plan_name, amount in zip(plan_names, amounts):
                    if amount is None:
                        continue
                    db.add(InsurerPlanCoverage(
                        insurer_id=insurer_id, plan_name=plan_name, coverage_label=label,
                        amount_text=str(amount), unit="만원", sort_order=sort_order,
                        source=_SOURCE, source_note=footnote, collected_at=_COLLECTED_AT,
                    ))
                sort_order += 1
            counts[insurer_code] = len(data)

        db.commit()
        for code, n in counts.items():
            print(f"{code}: 담보 {n}개 적재")
        all_codes = {i.code for i in db.query(Insurer).all()}
        missing = sorted(all_codes - set(counts))
        if missing:
            print(f"이 엑셀에 시트가 없어 건너뜀: {', '.join(missing)}")
        return counts
    finally:
        db.close()


if __name__ == "__main__":
    sys.exit(0 if run(Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_PATH) else 1)
