"""6개사를 같은 담보 항목 기준으로 나란히 비교한 표(보장비교 종합)를 적재한다.

seed_plan_coverage.py가 적재하는 "*_상세" 6개 시트는 보험사마다 원문 그대로의
담보명을 쓴다 — "이 항목이 저 보험사의 어느 항목과 같은 건지"를 사람이 다시
대조해야 한다. 이 스크립트는 사용자가 6개사를 같은 평가기준으로 재정리해서 준
"보장비교(종합)" 시트를 적재한다 — 같은 행에 6개사×등급의 값이 나란히 있어
바로 비교할 수 있다.

시트 구조(backend/data/source_files/insurer_plan_coverage_2026-08.xlsx의
"보장비교(종합)" 시트):

    ('평가기준', '카카오페이손해보험', None, None, '현대해상', None, None, ...)  <- 보험사 헤더(3칸씩)
    (None, '라이트', '베이직', '플러스', '실속', '표준', '고급', ...)              <- 등급 헤더
    ('사망 · 후유장해 (단위: 만원)', None, ..., None)                            <- 카테고리 헤더(라벨만, 나머지 None)
    ('상해사망보험금', 10000, 30000, 60000, ...)                                 <- 데이터 행(18칸 전부 값)
    ...
    ('참고: 예시 보험료 (원) ...', None, ..., None)                              <- 여기부터는 보험료 참고 섹션
    ...                                                                          (InsurerPremium에 이미 있어 건너뜀)

등급명이 보험사마다 이 시트에서만 다르게 적혀 있는 경우(예: 삼성은 "실속",
InsurerPlanCoverage/InsurerPremium은 "실속플랜")가 있어 _PLAN_ALIAS로
정규화한다 — 안 그러면 같은 보험사인데 표마다 등급명이 달라 화면에서 못 잇는다.

Run from ``backend``::

    python -m app.seed_comparison_metrics
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import openpyxl

from app.database import Base, SessionLocal, engine
from app import models  # noqa: F401  (모델 등록)
from app.models.kb import Insurer, InsurerComparisonMetric

DEFAULT_PATH = Path(__file__).resolve().parents[1] / "data" / "source_files" / "insurer_plan_coverage_2026-08.xlsx"

#: 시트의 보험사 헤더(row 0)에 적힌 이름 -> 보험사 코드, 그 보험사의 3개 등급 열이
#: 시작하는 열 인덱스(0-based, 'A'=0 기준). 각 보험사는 3칸을 차지한다.
_INSURER_COLUMNS: list[tuple[str, str, int]] = [
    ("카카오페이손해보험", "KAKAOPAY", 1),
    ("현대해상", "HYUNDAI", 4),
    ("KB손해보험", "KB", 7),
    ("삼성화재", "SAMSUNG", 10),
    ("DB손해보험", "DB", 13),
    ("메리츠화재", "MERITZ", 16),
]

#: 이 시트에서만 다르게 적힌 등급명을, 다른 두 표(InsurerPremium/InsurerPlanCoverage)와
#: 같은 표기로 맞춘다.
_PLAN_ALIAS: dict[str, dict[str, str]] = {
    "HYUNDAI": {"실속": "실속형", "표준": "표준형", "고급": "고급형"},
    "SAMSUNG": {"실속": "실속플랜", "표준": "표준플랜", "고급": "고급플랜"},
    "MERITZ": {"실속": "실속플랜", "표준(추천)": "추천플랜", "고급(보장큰)": "보장이큰플랜"},
}

_SOURCE = "각 사 다이렉트 가격공시/보험료 계산 화면(사용자가 직접 조회해 재정리)"
_COLLECTED_AT = date(2026, 8, 17)


def _normalize_plan(insurer_code: str, raw_plan: str) -> str:
    return _PLAN_ALIAS.get(insurer_code, {}).get(raw_plan, raw_plan)


def run(path: Path = DEFAULT_PATH) -> int:
    if not path.exists():
        raise SystemExit(f"{path} 가 없습니다. 담보 가입금액표 엑셀을 이 경로에 둔 뒤 다시 실행하세요.")

    Base.metadata.create_all(bind=engine)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = next((n for n in wb.sheetnames if n.startswith("보장비교")), None)
    if sheet_name is None:
        raise SystemExit("'보장비교(종합)' 시트를 찾지 못했습니다.")
    rows = list(wb[sheet_name].iter_rows(values_only=True))

    plan_header = rows[1]

    db = SessionLocal()
    try:
        code_to_id = {i.code: i.insurer_id for i in db.query(Insurer).all()}
        db.query(InsurerComparisonMetric).delete()

        # 각주(작성 안내 이후 문단)를 모아 표 전체 공통 각주로 남긴다.
        note_start = next(
            (i for i, r in enumerate(rows) if r[0] == "작성 안내"), None
        )
        source_note = " ".join(
            str(r[0]) for r in rows[note_start + 2:] if note_start is not None and r[0]
        ) if note_start is not None else None

        category = ""
        category_order = -1
        sort_order = 0
        inserted = 0

        for row in rows[2:]:
            label = row[0]
            if label is None:
                continue
            if isinstance(label, str) and label.startswith("참고:"):
                break  # 여기부터는 보험료 참고 섹션 — InsurerPremium에 이미 있어 건너뜀
            values = row[1:]
            if all(v is None for v in values):
                # 카테고리 헤더 행 — "사망 · 후유장해 (단위: 만원)"처럼 괄호 앞부분만 쓴다.
                category = str(label).split("(")[0].strip()
                category_order += 1
                sort_order = 0
                continue

            for insurer_name, insurer_code, col in _INSURER_COLUMNS:
                insurer_id = code_to_id.get(insurer_code)
                if insurer_id is None:
                    continue
                for offset in range(3):
                    raw_plan = plan_header[col + offset]
                    value = row[col + offset]
                    if raw_plan is None or value is None:
                        continue
                    plan_name = _normalize_plan(insurer_code, str(raw_plan))
                    db.add(InsurerComparisonMetric(
                        category=category, category_order=category_order,
                        metric_label=str(label), sort_order=sort_order,
                        insurer_id=insurer_id, plan_name=plan_name,
                        value_text=str(value).replace("\n", " "), unit="만원",
                        source=_SOURCE, source_note=source_note, collected_at=_COLLECTED_AT,
                    ))
                    inserted += 1
            sort_order += 1

        db.commit()
        print(f"보장비교 항목 {inserted}건 적재")
        return inserted
    finally:
        db.close()


if __name__ == "__main__":
    sys.exit(0 if run(Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_PATH) else 1)
